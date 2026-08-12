"""API helpers for the owner-information flow."""

import json
import os
import re
import threading
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager


URL_TRA_CUU_THUA_DAT = (
    "https://dla.mplis.gov.vn/dc/LamSachDuLieuAjax/"
    "GetThongKePhanLoaiThuaDatChiTiet"
)
URL_CAP_NHAT_PHAN_LOAI_THUA_DAT = (
    "https://dla.mplis.gov.vn/dc/LamSachDuLieuAjax/"
    "CapNhatThongKePhanLoaiThuaDatChiTiet"
)
REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
TIMEOUT = 120
REQUEST_DELAY_SECONDS = 0.15
DEFAULT_TINH_ID = 66
DEFAULT_PAGE_SIZE = 10
DEFAULT_SUB_LENGTH = 3000

THIEU_LOAI_DOI_TUONG_PATTERN = re.compile(
    r"^TINHHINHDANGKY\.(?P<tinh_hinh_dang_ky_id>\d+)"
    r"\|(?:(?:VOCHONG|HOGIADINH|NHOMNGUOI)\.[^|]+\|)*"
    r"CANHAN\.[^|]+\|loaiDoiTuongId$",
    re.IGNORECASE,
)


def _form_value(value: Any) -> str:
    """Convert a value to the representation expected by form-urlencoded APIs."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def build_tra_cuu_payload(
    so_thua: Any,
    so_to: Any,
    xa_id: Any,
    tinh_id: Any,
    *,
    so_phat_hanh: Any = "",
    ho_ten_chu: Any = "",
    phan_loai: Any = -1,
    loai_ban_ghi: Any = -1,
    loai_chu: Any = -1,
    tu_ngay: Any = "",
    den_ngay: Any = "",
    query: Any = "",
    huyen_id: Any = 0,
    start: int = 0,
    length: int = DEFAULT_PAGE_SIZE,
    export_ward: bool = False,
    sub_length: int = DEFAULT_SUB_LENGTH,
    sort_field: str = "_id",
    sort_direction: int = 1,
) -> dict[str, str]:
    """Build the form-urlencoded payload used to search owner information."""
    return {
        "traCuu[soThuTuThua]": _form_value(so_thua),
        "traCuu[soHieuToBanDo]": _form_value(so_to),
        "traCuu[soPhatHanh]": _form_value(so_phat_hanh),
        "traCuu[hoTenChu]": _form_value(ho_ten_chu),
        "traCuu[phanLoai]": _form_value(phan_loai),
        "traCuu[type]": _form_value(loai_ban_ghi),
        "traCuu[loaiChu]": _form_value(loai_chu),
        "traCuu[tuNgay]": _form_value(tu_ngay),
        "traCuu[denNgay]": _form_value(den_ngay),
        "traCuu[query]": _form_value(query),
        "traCuu[xaId]": _form_value(xa_id),
        "traCuu[huyenId]": _form_value(huyen_id),
        "traCuu[tinhId]": _form_value(tinh_id),
        "start": _form_value(start),
        "length": _form_value(length),
        "exportWard": _form_value(export_ward),
        "subLength": _form_value(sub_length),
        "sort[Field]": _form_value(sort_field),
        "sort[Direction]": _form_value(sort_direction),
    }


def tra_cuu_thua_dat(
    session: requests.Session,
    so_thua: Any,
    so_to: Any,
    xa_id: Any,
    tinh_id: Any,
    **payload_options: Any,
) -> dict[str, Any]:
    """Search parcels using an authenticated MPLIS session."""
    payload = build_tra_cuu_payload(
        so_thua=so_thua,
        so_to=so_to,
        xa_id=xa_id,
        tinh_id=tinh_id,
        **payload_options,
    )
    response = session.post(
        URL_TRA_CUU_THUA_DAT,
        data=payload,
        headers={
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        },
        timeout=TIMEOUT,
    )
    response.raise_for_status()

    try:
        result = response.json()
    except requests.JSONDecodeError as exc:
        raise RuntimeError(
            "TRA CUU THUA DAT khong tra JSON "
            f"(HTTP {response.status_code}): {response.text[:500]}"
        ) from exc

    if not isinstance(result, dict):
        raise RuntimeError("TRA CUU THUA DAT tra ve du lieu khong hop le.")
    return result


def lay_tinh_hinh_dang_ky_ids_thieu_loai_doi_tuong(
    response_data: dict[str, Any],
) -> list[int]:
    """Return registration IDs whose personal owner has no object type."""
    if response_data.get("success") is not True:
        raise RuntimeError(
            "TRA CUU THUA DAT that bai: " + str(response_data)[:500]
        )

    rows = response_data.get("data")
    if not isinstance(rows, list):
        raise RuntimeError("TRA CUU THUA DAT khong co danh sach data hop le.")

    result: list[int] = []
    seen: set[int] = set()

    for row in rows:
        if not isinstance(row, dict):
            continue

        try:
            row_id = int(row["tinhHinhDangKyId"])
        except (KeyError, TypeError, ValueError):
            continue

        missing_fields = row.get("thongTinDangKyChuaDapUngNhom1")
        if not isinstance(missing_fields, list):
            continue

        for missing_field in missing_fields:
            if not isinstance(missing_field, str):
                continue
            match = THIEU_LOAI_DOI_TUONG_PATTERN.fullmatch(missing_field.strip())
            if not match:
                continue
            if int(match.group("tinh_hinh_dang_ky_id")) != row_id:
                continue
            if row_id not in seen:
                seen.add(row_id)
                result.append(row_id)
            break

    return result


def tra_cuu_ids_thieu_loai_doi_tuong(
    session: requests.Session,
    so_thua: Any,
    so_to: Any,
    xa_id: Any,
    tinh_id: Any,
    **payload_options: Any,
) -> list[int]:
    """Search and select only registration IDs missing owner object type."""
    response_data = tra_cuu_thua_dat(
        session=session,
        so_thua=so_thua,
        so_to=so_to,
        xa_id=xa_id,
        tinh_id=tinh_id,
        **payload_options,
    )
    return lay_tinh_hinh_dang_ky_ids_thieu_loai_doi_tuong(response_data)


def build_cap_nhat_loai_doi_tuong_payload(
    row: dict[str, Any],
    loai_doi_tuong_id: Any = 22,
) -> dict[str, Any]:
    """Build one update payload from an eligible parcel-search result row."""
    record_id = str(row.get("id") or "").strip()
    if not record_id:
        raise ValueError("Ban ghi thieu id de cap nhat.")

    try:
        registration_id = int(row["tinhHinhDangKyId"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("Ban ghi co tinhHinhDangKyId khong hop le.") from exc
    if registration_id <= 0:
        raise ValueError("tinhHinhDangKyId phai lon hon 0.")

    object_type = _form_value(loai_doi_tuong_id)
    if not object_type:
        raise ValueError("loaiDoiTuongId khong duoc rong.")

    missing_fields = row.get("thongTinDangKyChuaDapUngNhom1")
    if not isinstance(missing_fields, list):
        raise ValueError("Ban ghi khong co danh sach thong tin nhom 1 bi thieu.")

    updates: dict[str, str] = {}
    for missing_field in missing_fields:
        if not isinstance(missing_field, str):
            continue
        field_key = missing_field.strip()
        match = THIEU_LOAI_DOI_TUONG_PATTERN.fullmatch(field_key)
        if not match:
            continue
        if int(match.group("tinh_hinh_dang_ky_id")) == registration_id:
            updates[field_key] = object_type

    if not updates:
        raise ValueError("Ban ghi khong thieu loaiDoiTuongId cua ca nhan.")

    return {
        "id": record_id,
        "tinhHinhDangKyId": registration_id,
        "data": json.dumps(updates, ensure_ascii=False, separators=(",", ":")),
    }


def build_cap_nhat_payloads_from_response(
    response_data: dict[str, Any],
    loai_doi_tuong_id: Any = 22,
) -> list[dict[str, Any]]:
    """Build update payloads only for rows missing personal-owner object type."""
    if response_data.get("success") is not True:
        raise RuntimeError(
            "TRA CUU THUA DAT that bai: " + str(response_data)[:500]
        )
    rows = response_data.get("data")
    if not isinstance(rows, list):
        raise RuntimeError("TRA CUU THUA DAT khong co danh sach data hop le.")

    payloads: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            payloads.append(
                build_cap_nhat_loai_doi_tuong_payload(row, loai_doi_tuong_id)
            )
        except ValueError:
            continue
    return payloads


def cap_nhat_loai_doi_tuong(
    session: requests.Session,
    payload: dict[str, Any],
) -> dict[str, Any]:
    """Submit one owner object-type update."""
    response = session.post(
        URL_CAP_NHAT_PHAN_LOAI_THUA_DAT,
        json=payload,
        headers={"Content-Type": "application/json; charset=UTF-8"},
        timeout=TIMEOUT,
    )
    response.raise_for_status()

    try:
        result = response.json()
    except requests.JSONDecodeError as exc:
        raise RuntimeError(
            "CAP NHAT LOAI DOI TUONG khong tra JSON "
            f"(HTTP {response.status_code}): {response.text[:500]}"
        ) from exc

    if not isinstance(result, dict):
        raise RuntimeError("CAP NHAT LOAI DOI TUONG tra ve du lieu khong hop le.")
    if result.get("success") is False:
        raise RuntimeError("CAP NHAT LOAI DOI TUONG that bai: " + str(result)[:500])
    return result


def tra_cuu_va_cap_nhat_loai_doi_tuong(
    session: requests.Session,
    so_thua: Any,
    so_to: Any,
    xa_id: Any,
    tinh_id: Any,
    *,
    loai_doi_tuong_id: Any = 22,
    **payload_options: Any,
) -> list[dict[str, Any]]:
    """Run search, filter eligible rows, then submit their updates."""
    search_result = tra_cuu_thua_dat(
        session=session,
        so_thua=so_thua,
        so_to=so_to,
        xa_id=xa_id,
        tinh_id=tinh_id,
        **payload_options,
    )
    payloads = build_cap_nhat_payloads_from_response(
        search_result, loai_doi_tuong_id=loai_doi_tuong_id
    )
    return [cap_nhat_loai_doi_tuong(session, payload) for payload in payloads]


# ============================ LOGIN / SESSION ============================


def lay_token_tu_trang(driver: webdriver.Chrome) -> str:
    script = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return str(driver.execute_script(script) or "").strip()


class MplisLoginClient:
    def __init__(self, log_fn):
        self.log = log_fn
        self.driver: webdriver.Chrome | None = None
        self.session: requests.Session | None = None

    def open_browser_and_fill_login(self, username: str, password: str) -> None:
        options = Options()
        options.add_argument("--start-maximized")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.get(REFERER_URL)
        time.sleep(2)

        inputs = self.driver.find_elements(By.CSS_SELECTOR, "input")
        user_box = None
        password_box = None
        for input_element in inputs:
            input_type = (input_element.get_attribute("type") or "").lower()
            if user_box is None and input_type in {"text", "email"}:
                user_box = input_element
            if password_box is None and input_type == "password":
                password_box = input_element

        if user_box is None or password_box is None:
            self.log("Không tự nhận diện được form. Hãy đăng nhập trực tiếp trên Chrome.")
            return

        user_box.clear()
        user_box.send_keys(username)
        password_box.clear()
        password_box.send_keys(password)
        password_box.send_keys(Keys.ENTER)

    def build_session_from_browser(self) -> None:
        if self.driver is None:
            raise RuntimeError("Chưa mở Chrome đăng nhập.")

        token = lay_token_tu_trang(self.driver)
        if not token:
            raise RuntimeError(
                "Không lấy được token. Kiểm tra đã đăng nhập và trang MPLIS đã tải xong."
            )

        session = requests.Session()
        user_agent = self.driver.execute_script("return navigator.userAgent;")
        session.headers.update(
            {
                "User-Agent": user_agent,
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
                "Origin": "https://dla.mplis.gov.vn",
                "Referer": REFERER_URL,
                "__requestverificationtoken": token,
                "__RequestVerificationToken": token,
                "RequestVerificationToken": token,
            }
        )

        for cookie in self.driver.get_cookies():
            session.cookies.set(
                name=cookie["name"],
                value=cookie["value"],
                domain=cookie.get("domain"),
                path=cookie.get("path", "/"),
            )

        self.session = session

    def require_session(self) -> requests.Session:
        if self.session is None:
            raise RuntimeError("Chưa lấy session đăng nhập MPLIS.")
        return self.session

    def close_browser(self) -> None:
        if self.driver is not None:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

    def close(self) -> None:
        self.close_browser()
        if self.session is not None:
            self.session.close()
            self.session = None


# ============================ EXCEL ============================


def chuan_hoa_ten_cot(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return "".join(char for char in text if char.isalnum())


def chuan_hoa_gia_tri_excel(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def tim_cot_to_thua(worksheet) -> tuple[int, int, int]:
    aliases_so_to = {"soto", "sotobando", "sohieutobando", "tobando", "to"}
    aliases_so_thua = {"sothua", "sothututhua", "thuadat", "thua"}

    for row_index in range(1, min(10, worksheet.max_row) + 1):
        so_to_column = None
        so_thua_column = None
        for column_index in range(1, worksheet.max_column + 1):
            column_name = chuan_hoa_ten_cot(
                worksheet.cell(row=row_index, column=column_index).value
            )
            if column_name in aliases_so_to and so_to_column is None:
                so_to_column = column_index
            if column_name in aliases_so_thua and so_thua_column is None:
                so_thua_column = column_index
        if so_to_column and so_thua_column:
            return row_index, so_to_column, so_thua_column

    raise ValueError(
        "Không tìm thấy cột Số tờ và Số thửa trong 10 hàng đầu của Excel."
    )


def doc_danh_sach_thua(file_path: str) -> list[dict[str, Any]]:
    extension = Path(file_path).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("Chỉ hỗ trợ file Excel .xlsx hoặc .xlsm.")

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError("File Excel không hợp lệ hoặc đã bị hỏng.") from exc

    try:
        worksheet = workbook.active
        header_row, so_to_column, so_thua_column = tim_cot_to_thua(worksheet)
        parcels: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()

        for row_index in range(header_row + 1, worksheet.max_row + 1):
            so_to = chuan_hoa_gia_tri_excel(
                worksheet.cell(row=row_index, column=so_to_column).value
            )
            so_thua = chuan_hoa_gia_tri_excel(
                worksheet.cell(row=row_index, column=so_thua_column).value
            )
            if not so_to and not so_thua:
                continue

            key = (so_to, so_thua)
            if key in seen:
                continue
            seen.add(key)
            parcels.append(
                {"excel_row": row_index, "so_to": so_to, "so_thua": so_thua}
            )

        if not parcels:
            raise ValueError("Excel không có dữ liệu số tờ/số thửa để xử lý.")
        return parcels
    finally:
        workbook.close()


class ExcelResultWriter:
    HEADERS = [
        "STT",
        "Số tờ",
        "Số thửa",
        "Số bản ghi tìm thấy",
        "tinhHinhDangKyIds",
        "Số bản ghi cập nhật",
        "Kết quả",
        "Chi tiết",
    ]

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "KetQua"
        self.worksheet.append(self.HEADERS)
        self.row_count = 0

        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in self.worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, width in enumerate([8, 12, 12, 20, 28, 22, 22, 70], start=1):
            self.worksheet.column_dimensions[get_column_letter(index)].width = width
        self.worksheet.freeze_panes = "A2"
        self.worksheet.auto_filter.ref = "A1:H1"

    def append(
        self,
        so_to: str,
        so_thua: str,
        found_count: int,
        registration_ids: list[int],
        updated_count: int,
        result: str,
        detail: str,
    ) -> None:
        self.row_count += 1
        self.worksheet.append(
            [
                self.row_count,
                so_to,
                so_thua,
                found_count,
                "; ".join(str(value) for value in registration_ids),
                updated_count,
                result,
                detail,
            ]
        )
        self.worksheet.cell(self.row_count + 1, 8).alignment = Alignment(wrap_text=True)

    def save(self) -> None:
        output_parent = Path(self.output_path).resolve().parent
        output_parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)


# ============================ UI ============================


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Cập nhật loại đối tượng chủ MPLIS")
        self.root.geometry("920x680")
        self.root.minsize(820, 600)

        self.client = MplisLoginClient(self.log)
        self.running = False
        self.stop_flag = False
        self.var_input_file = tk.StringVar()
        self.var_output_file = tk.StringVar()

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self) -> None:
        login_frame = ttk.LabelFrame(self.root, text="Đăng nhập MPLIS", padding=10)
        login_frame.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Label(login_frame, text="Username:").grid(row=0, column=0, sticky="w")
        self.ent_user = ttk.Entry(login_frame, width=30)
        self.ent_user.grid(row=0, column=1, sticky="ew", padx=5, pady=3)

        ttk.Label(login_frame, text="Password:").grid(row=0, column=2, sticky="w")
        self.ent_pass = ttk.Entry(login_frame, width=30, show="*")
        self.ent_pass.grid(row=0, column=3, sticky="ew", padx=5, pady=3)

        ttk.Label(login_frame, text="Mã xã (xaId):").grid(row=1, column=0, sticky="w")
        self.ent_xa_id = ttk.Entry(login_frame, width=20)
        self.ent_xa_id.grid(row=1, column=1, sticky="w", padx=5, pady=3)
        login_frame.columnconfigure(1, weight=1)
        login_frame.columnconfigure(3, weight=1)

        file_frame = ttk.LabelFrame(self.root, text="File Excel", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(file_frame, text="File đầu vào:").grid(row=0, column=0, sticky="w")
        ttk.Entry(file_frame, textvariable=self.var_input_file).grid(
            row=0, column=1, sticky="ew", padx=5, pady=3
        )
        ttk.Button(file_frame, text="Duyệt file...", command=self.chon_file_input).grid(
            row=0, column=2, padx=5, pady=3
        )

        ttk.Label(file_frame, text="File kết quả:").grid(row=1, column=0, sticky="w")
        ttk.Entry(file_frame, textvariable=self.var_output_file).grid(
            row=1, column=1, sticky="ew", padx=5, pady=3
        )
        ttk.Button(file_frame, text="Chọn nơi lưu...", command=self.chon_file_output).grid(
            row=1, column=2, padx=5, pady=3
        )
        file_frame.columnconfigure(1, weight=1)

        button_frame = ttk.Frame(self.root, padding=(10, 5))
        button_frame.pack(fill="x")
        self.btn_login = ttk.Button(
            button_frame, text="1. Mở Chrome đăng nhập", command=self.mo_chrome
        )
        self.btn_login.pack(side="left", padx=5)
        self.btn_session = ttk.Button(
            button_frame,
            text="2. Lấy session",
            command=self.lay_session,
            state="disabled",
        )
        self.btn_session.pack(side="left", padx=5)
        self.btn_run = ttk.Button(
            button_frame,
            text="3. Xử lý Excel",
            command=self.chay,
            state="disabled",
        )
        self.btn_run.pack(side="left", padx=5)
        self.btn_stop = ttk.Button(
            button_frame, text="Dừng", command=self.dung, state="disabled"
        )
        self.btn_stop.pack(side="left", padx=5)

        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=(5, 0))
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill="x", padx=10, pady=(3, 0))
        self.lbl_status = ttk.Label(status_frame, text="Chưa chạy")
        self.lbl_status.pack(side="left")
        self.lbl_count = ttk.Label(status_frame, text="")
        self.lbl_count.pack(side="right")

        log_frame = ttk.LabelFrame(self.root, text="Nhật ký xử lý", padding=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=8)
        self.txt_log = tk.Text(log_frame, wrap="word", height=24)
        self.txt_log.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(log_frame, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=scrollbar.set)

    def log(self, message: str) -> None:
        def append() -> None:
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.txt_log.insert("end", f"{timestamp}  {message}\n")
            self.txt_log.see("end")

        self.root.after(0, append)

    def set_status(self, message: str) -> None:
        self.root.after(0, lambda: self.lbl_status.config(text=message))

    def set_progress(self, value: int, maximum: int) -> None:
        def update() -> None:
            self.progress.config(maximum=max(maximum, 1), value=value)
            self.lbl_count.config(text=f"{value} / {maximum}")

        self.root.after(0, update)

    def chon_file_input(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Chọn Excel có cột Số tờ và Số thửa",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled", "*.xlsm"),
                ("Tất cả file", "*.*"),
            ],
        )
        if not file_path:
            return
        self.var_input_file.set(file_path)
        input_path = Path(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.var_output_file.set(
            str(input_path.with_name(f"ket_qua_cap_nhat_loai_chu_{timestamp}.xlsx"))
        )

    def chon_file_output(self) -> None:
        initial_name = Path(
            self.var_output_file.get() or "ket_qua_cap_nhat_loai_chu.xlsx"
        ).name
        file_path = filedialog.asksaveasfilename(
            title="Chọn nơi lưu file kết quả",
            defaultextension=".xlsx",
            initialfile=initial_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if file_path:
            self.var_output_file.set(file_path)

    def mo_chrome(self) -> None:
        username = self.ent_user.get().strip()
        password = self.ent_pass.get()
        if not username or not password:
            messagebox.showwarning("Thiếu thông tin", "Nhập username và password.")
            return

        self.btn_login.config(state="disabled")

        def work() -> None:
            try:
                self.log("Đang mở Chrome...")
                self.client.open_browser_and_fill_login(username, password)
                self.log("Hoàn tất đăng nhập trên Chrome rồi bấm '2. Lấy session'.")
                self.root.after(0, lambda: self.btn_session.config(state="normal"))
            except Exception as exc:
                self.log(f"Lỗi mở Chrome: {exc}")
                self.root.after(0, lambda: self.btn_login.config(state="normal"))

        threading.Thread(target=work, daemon=True).start()

    def lay_session(self) -> None:
        self.btn_session.config(state="disabled")

        def work() -> None:
            try:
                self.client.build_session_from_browser()
                self.client.close_browser()
                self.log("Đã lấy cookie và token; Chrome đã đóng.")
                self.root.after(0, lambda: self.btn_run.config(state="normal"))
            except Exception as exc:
                self.log(f"Lỗi lấy session: {exc}")
                self.root.after(0, lambda: self.btn_session.config(state="normal"))

        threading.Thread(target=work, daemon=True).start()

    def dung(self) -> None:
        self.stop_flag = True
        self.log("Đã yêu cầu dừng; sẽ dừng sau request hiện tại.")

    def chay(self) -> None:
        if self.running:
            return

        xa_id = self.ent_xa_id.get().strip()
        input_file = self.var_input_file.get().strip()
        output_file = self.var_output_file.get().strip()

        if not xa_id.isdigit():
            messagebox.showwarning("Sai mã xã", "Mã xã phải là số nguyên.")
            return
        if not os.path.isfile(input_file):
            messagebox.showwarning("Thiếu file", "Chọn file Excel đầu vào hợp lệ.")
            return
        if not output_file:
            messagebox.showwarning("Thiếu file", "Chọn đường dẫn file kết quả.")
            return
        if Path(input_file).resolve() == Path(output_file).resolve():
            messagebox.showwarning("Sai đường dẫn", "File kết quả không được trùng file đầu vào.")
            return
        try:
            self.client.require_session()
            parcels = doc_danh_sach_thua(input_file)
        except Exception as exc:
            messagebox.showerror("Không thể chạy", str(exc))
            return

        if not messagebox.askyesno(
            "Xác nhận cập nhật",
            f"Sẽ xử lý {len(parcels)} thửa đất tại xã {xa_id}.\n"
            "Các cá nhân đang thiếu loại đối tượng sẽ được gán mã 22.\n"
            "Tiếp tục?",
        ):
            return

        self.running = True
        self.stop_flag = False
        self.btn_run.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.set_progress(0, len(parcels))
        threading.Thread(
            target=self._run_batch,
            args=(xa_id, parcels, output_file),
            daemon=True,
        ).start()

    def _run_batch(
        self,
        xa_id: str,
        parcels: list[dict[str, Any]],
        output_file: str,
    ) -> None:
        writer = ExcelResultWriter(output_file)
        updated_total = 0
        skipped_total = 0
        error_total = 0
        processed = 0

        try:
            session = self.client.require_session()
            for index, parcel in enumerate(parcels, start=1):
                if self.stop_flag:
                    self.log("Đã dừng theo yêu cầu.")
                    break

                so_to = parcel["so_to"]
                so_thua = parcel["so_thua"]
                self.set_status(f"Đang xử lý tờ {so_to or '?'} - thửa {so_thua or '?'}")
                self.log(
                    f"[{index}/{len(parcels)}] Dòng {parcel['excel_row']}: "
                    f"tờ {so_to}, thửa {so_thua}"
                )

                found_count = 0
                registration_ids: list[int] = []
                updated_count = 0
                result_text = ""
                detail = ""

                try:
                    if not so_to or not so_thua:
                        raise ValueError("Thiếu số tờ hoặc số thửa.")

                    search_result = tra_cuu_thua_dat(
                        session=session,
                        so_thua=so_thua,
                        so_to=so_to,
                        xa_id=xa_id,
                        tinh_id=DEFAULT_TINH_ID,
                    )
                    rows = search_result.get("data")
                    if not isinstance(rows, list):
                        raise RuntimeError("Response tra cứu không có danh sách data hợp lệ.")
                    found_count = len(rows)
                    payloads = build_cap_nhat_payloads_from_response(search_result)
                    registration_ids = [
                        payload["tinhHinhDangKyId"] for payload in payloads
                    ]

                    if not rows:
                        result_text = "KHÔNG TÌM THẤY"
                        detail = "Không có bản ghi khớp tờ/thửa."
                        skipped_total += 1
                    elif not payloads:
                        result_text = "BỎ QUA"
                        detail = "Không có cá nhân thiếu loaiDoiTuongId."
                        skipped_total += 1
                    else:
                        messages = []
                        for payload in payloads:
                            cap_nhat_loai_doi_tuong(session, payload)
                            updated_count += 1
                            updated_total += 1
                            messages.append(
                                f"Đã cập nhật THĐK {payload['tinhHinhDangKyId']}"
                            )
                        result_text = "ĐÃ CẬP NHẬT"
                        detail = "; ".join(messages)

                    self.log(f"   {result_text}: {detail}")
                except Exception as exc:
                    result_text = "LỖI"
                    detail = str(exc)
                    error_total += 1
                    self.log(f"   LỖI: {detail}")

                writer.append(
                    so_to=so_to,
                    so_thua=so_thua,
                    found_count=found_count,
                    registration_ids=registration_ids,
                    updated_count=updated_count,
                    result=result_text,
                    detail=detail,
                )
                processed = index
                if index % 5 == 0:
                    writer.save()
                self.set_progress(index, len(parcels))
                time.sleep(REQUEST_DELAY_SECONDS)
        finally:
            try:
                writer.save()
                self.log(f"Đã lưu file kết quả: {output_file}")
            except Exception as exc:
                self.log(f"Lỗi lưu file kết quả: {exc}")

            self.running = False
            self.set_status(
                f"Hoàn tất {processed}/{len(parcels)}: "
                f"{updated_total} cập nhật, {skipped_total} bỏ qua, {error_total} lỗi"
            )
            self.root.after(0, lambda: self.btn_run.config(state="normal"))
            self.root.after(0, lambda: self.btn_stop.config(state="disabled"))

    def on_close(self) -> None:
        if self.running and not messagebox.askyesno(
            "Đang xử lý", "Tiến trình đang chạy. Thoát ứng dụng?"
        ):
            return
        self.stop_flag = True
        self.client.close()
        self.root.destroy()


def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
