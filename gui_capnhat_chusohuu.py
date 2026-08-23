# -*- coding: utf-8 -*-
"""
GUI (CustomTkinter) cập nhật hàng loạt thông tin Chủ sở hữu (Tổ chức) trên MPLIS.
- Đăng nhập bằng Selenium (mở Chrome, tự điền user/pass, người dùng xử lý captcha/OTP nếu
  có rồi bấm "Đã đăng nhập xong") để lấy Cookie + __RequestVerificationToken.
- Duyệt file Excel: mỗi dòng có cột "GCN" (Số phát hành) để tra cứu ra ToChuc hiện tại
  (qua AdvancedSearchGiayChungNhan), cùng các cột thông tin tổ chức muốn cập nhật - cột nào
  có giá trị thì ghi đè, cột nào để trống thì giữ nguyên dữ liệu gốc lấy từ MPLIS.
- Mặc định chỉ KIỂM TRA (build payload, không gửi lên MPLIS). Tick "Gửi update thật lên MPLIS"
  mới thực sự gọi ChuSoHuuAjax/UpdateCaNhan + UpdateToChuc.
- Trước khi chạy sẽ tự kiểm tra các cột mã/số định danh (Số định danh tổ chức/người đại diện,
  Mã doanh nghiệp, Mã số thuế, Số giấy tờ GPKD): nếu ô nào trong file Excel gốc đang ở định dạng
  Số (không phải Text) thì cảnh báo, vì số 0 ở đầu (nếu có) đã bị Excel làm rụng mất trước khi
  đọc file - không khôi phục lại được, phải định dạng lại cột đó là Text trong Excel rồi nhập lại.

Cột Excel:
  BẮT BUỘC: "GCN" (Số phát hành, vd "DG 781336")
  TÙY CHỌN: "Tỉnh" (mặc định 66 nếu để trống/không có cột)
  TÙY CHỌN (để trống = giữ nguyên):
    "Tên tổ chức", "Địa chỉ tổ chức", "Số định danh tổ chức",
    "Mã doanh nghiệp", "Mã số thuế",
    "Số giấy tờ (GPKD)", "Ngày cấp GPKD" (dd/mm/yyyy),
    "Họ tên người đại diện", "Số định danh người đại diện",
    "Ngày sinh người đại diện" (dd/mm/yyyy), "Địa chỉ người đại diện"

Cài đặt: pip install customtkinter selenium webdriver-manager requests pandas openpyxl
Chạy: python gui_capnhat_chusohuu.py
"""

import os
import re
import copy
import threading
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
import requests

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import tracuu_giaychungnhan as gcn
import capnhat_chusohuu as cnh


# ============================ CẤU HÌNH ============================

REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
TINH_ID_MAC_DINH = 66

COL_GCN = "GCN"
COL_TINH = "Tỉnh"
REQUIRED_COLS = [COL_GCN]

# Cột Excel -> field ghi đè trên object ToChuc (cấp tổ chức, không gồm người đại diện)
OVERRIDE_COL_MAP = {
    "Tên tổ chức": "tenToChuc",
    "Địa chỉ tổ chức": "diaChi",
    "Số định danh tổ chức": "maSoDinhDanh",
    "Mã doanh nghiệp": "maDoanhNghiep",
    "Mã số thuế": "maSoThue",
}
OVERRIDE_DATE_COL_MAP = {}  # cột dạng ngày dd/mm/yyyy cấp tổ chức (hiện chưa có)

# Cột Excel -> field PHẲNG (không tiền tố) của người đại diện - dùng để: (1) gọi UpdateCaNhan
# sửa đúng bản ghi CaNhan gốc, và (2) ghi đè bản sao "NguoiDaiDien.xxx" trong payload ToChuc.
DAI_DIEN_COL_MAP = {
    "Họ tên người đại diện": "hoTen",
    "Số định danh người đại diện": "maSoDinhDanh",
    "Địa chỉ người đại diện": "diaChi",
}
DAI_DIEN_DATE_COL_MAP = {
    "Ngày sinh người đại diện": "ngaySinh",
}
COL_SO_GIAY_TO = "Số giấy tờ (GPKD)"
COL_NGAY_CAP_GIAY_TO = "Ngày cấp GPKD"

OPTIONAL_COLS = (
    [COL_TINH]
    + list(OVERRIDE_COL_MAP.keys())
    + list(OVERRIDE_DATE_COL_MAP.keys())
    + list(DAI_DIEN_COL_MAP.keys())
    + list(DAI_DIEN_DATE_COL_MAP.keys())
    + [COL_SO_GIAY_TO, COL_NGAY_CAP_GIAY_TO]
)

# Cột chứa mã/số định danh - nếu ô Excel gốc được lưu dưới dạng Số (không phải Text) thì số 0
# ở đầu (vd "0301234567") đã bị Excel làm rụng mất TRƯỚC KHI file được đọc - không có cách nào
# khôi phục lại từ dữ liệu đã đọc được nữa, phải sửa lại ở file Excel gốc (định dạng cột là
# Text rồi nhập lại, hoặc gõ thêm dấu ' trước số). Các cột này cần cảnh báo sớm cho người dùng.
COT_CO_THE_MAT_SO_0_DAU = [
    "Số định danh tổ chức",
    "Số định danh người đại diện",
    "Mã doanh nghiệp",
    "Mã số thuế",
    "Số giấy tờ (GPKD)",
]


# ============================ HELPER ============================

def clean_cell(v):
    """Chuẩn hóa giá trị đọc từ Excel: bỏ NaN, bỏ .0 của số, strip khoảng trắng."""
    if v is None:
        return ""
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        if v == int(v):
            return str(int(v))
        return str(v)
    s = str(v).strip()
    if s.lower() == "nan":
        return ""
    return s


def ngay_vn_sang_iso(ddmmyyyy):
    """Đổi ngày nhập dd/mm/yyyy (hiểu là 00:00 giờ VN) sang chuỗi ISO UTC."""
    s = (ddmmyyyy or "").strip()
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$", s)
    if not m:
        raise ValueError(f"Ngày không đúng định dạng dd/mm/yyyy: '{ddmmyyyy}'")
    d, mo, y = map(int, m.groups())
    dt_vn = datetime(y, mo, d)
    dt_utc = dt_vn - timedelta(hours=7)
    return dt_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")


def lay_token_tu_trang(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def build_overrides_tu_dong(row):
    """Chỉ đưa cột nào thực sự có giá trị (sau khi clean_cell) vào overrides - cột trống giữ
    nguyên dữ liệu gốc lấy từ MPLIS. Trả về (overrides_to_chuc, dai_dien_overrides, ngay_cap, so_giay_to):
    - overrides_to_chuc: field cấp Tổ chức (kèm tiền tố "NguoiDaiDien." cho bản sao gắn trong đó)
    - dai_dien_overrides: field PHẲNG của người đại diện, dùng gọi UpdateCaNhan riêng"""
    overrides = {}
    for cot, field in OVERRIDE_COL_MAP.items():
        gia_tri = clean_cell(row.get(cot))
        if gia_tri:
            overrides[field] = gia_tri

    for cot, field in OVERRIDE_DATE_COL_MAP.items():
        gia_tri = clean_cell(row.get(cot))
        if gia_tri:
            overrides[field] = ngay_vn_sang_iso(gia_tri)

    dai_dien_overrides = {}
    for cot, field in DAI_DIEN_COL_MAP.items():
        gia_tri = clean_cell(row.get(cot))
        if gia_tri:
            dai_dien_overrides[field] = gia_tri
            overrides[f"NguoiDaiDien.{field}"] = gia_tri

    for cot, field in DAI_DIEN_DATE_COL_MAP.items():
        gia_tri = clean_cell(row.get(cot))
        if gia_tri:
            gia_tri_iso = ngay_vn_sang_iso(gia_tri)
            dai_dien_overrides[field] = gia_tri_iso
            overrides[f"NguoiDaiDien.{field}"] = gia_tri_iso

    so_giay_to = clean_cell(row.get(COL_SO_GIAY_TO)) or None
    ngay_cap_raw = clean_cell(row.get(COL_NGAY_CAP_GIAY_TO))
    ngay_cap = ngay_vn_sang_iso(ngay_cap_raw) if ngay_cap_raw else None

    return overrides, dai_dien_overrides, ngay_cap, so_giay_to


def tim_o_dinh_dang_so(path, cot_can_kiem):
    """Đọc trực tiếp bằng openpyxl (giữ được kiểu dữ liệu gốc của ô, khác pandas.read_excel đã
    làm mất thông tin này) để tìm các ô thuộc cot_can_kiem được lưu dưới dạng Số (cell.data_type
    == 'n'). Đây là dấu hiệu số 0 ở đầu (nếu có) đã bị Excel làm rụng mất khi lưu file - không
    khôi phục lại được từ dữ liệu đã đọc, chỉ có thể cảnh báo để người dùng tự kiểm tra/sửa lại
    file Excel gốc (định dạng cột là Text rồi nhập lại). Trả về list (dòng Excel, tên cột, giá
    trị đang đọc được)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        hang_dau = next(ws.iter_rows(min_row=1, max_row=1))
        header = [str(o.value).strip() if o.value is not None else "" for o in hang_dau]
        cot_idx = {ten: i for i, ten in enumerate(header) if ten in cot_can_kiem}
        if not cot_idx:
            return []

        canh_bao = []
        for o in ws.iter_rows(min_row=2):
            for ten_cot, idx in cot_idx.items():
                if idx >= len(o):
                    continue
                cell = o[idx]
                if cell.data_type == "n" and cell.value not in (None, ""):
                    canh_bao.append((cell.row, ten_cot, cell.value))
        return canh_bao
    finally:
        wb.close()


def ghi_de_giay_to_bo_sung(to_chuc, ngay_cap=None, so_giay_to=None):
    """Sửa ngày cấp / số giấy tờ (GPKD) trong ListGiayToBoSung[0] nếu có giá trị mới;
    field nào không truyền thì giữ nguyên. Trả về bản sao, không sửa object gốc."""
    if ngay_cap is None and so_giay_to is None:
        return to_chuc
    to_chuc = copy.deepcopy(to_chuc)
    danh_sach = to_chuc.get("ListGiayToBoSung") or []
    if danh_sach:
        if ngay_cap is not None:
            danh_sach[0]["ngayCap"] = ngay_cap
        if so_giay_to is not None:
            danh_sach[0]["soGiayTo"] = so_giay_to
    return to_chuc


def xu_ly_1_dong(session, so_phat_hanh, tinh_id, overrides, dai_dien_overrides, ngay_cap, so_giay_to, gui_that):
    """Tra GCN -> lấy ToChuc -> (nếu có sửa người đại diện) UpdateCaNhan TRƯỚC -> áp overrides
    Tổ chức -> build + (tuỳ chọn) gửi UpdateToChuc. Trả về dict:
    {ca_nhan_payload, ca_nhan_result, to_chuc_payload, to_chuc_result} - *_result=None nếu
    chỉ xem trước (gui_that=False) hoặc bước đó không áp dụng (không sửa người đại diện)."""
    js_gcn = gcn.tra_cuu_giay_chung_nhan(session, so_phat_hanh, tinh_id)
    rows = js_gcn.get("data") or []
    if not rows:
        raise RuntimeError(f"Không tìm thấy GCN nào khớp Số phát hành: {so_phat_hanh}")

    to_chuc = None
    for row in rows:
        to_chuc = gcn.lay_to_chuc_tu_gcn_row(row)
        if to_chuc:
            break
    if not to_chuc:
        raise RuntimeError("Chủ sở hữu không phải Tổ chức (có thể là Cá nhân/Hộ gia đình).")

    ket_qua = {"ca_nhan_payload": None, "ca_nhan_result": None, "to_chuc_payload": None, "to_chuc_result": None}

    if dai_dien_overrides:
        ca_nhan = to_chuc.get("NguoiDaiDien")
        if not ca_nhan:
            raise RuntimeError("Có sửa thông tin người đại diện nhưng ToChuc không có NguoiDaiDien.")
        ca_nhan_payload = cnh.build_update_ca_nhan_payload(ca_nhan, **dai_dien_overrides)
        ket_qua["ca_nhan_payload"] = ca_nhan_payload
        if gui_that:
            ket_qua["ca_nhan_result"] = cnh.update_ca_nhan(session, ca_nhan_payload)

    to_chuc = ghi_de_giay_to_bo_sung(to_chuc, ngay_cap, so_giay_to)
    to_chuc_payload = cnh.build_update_to_chuc_payload(to_chuc, **overrides)
    ket_qua["to_chuc_payload"] = to_chuc_payload

    if gui_that:
        ket_qua["to_chuc_result"] = cnh.update_to_chuc(session, to_chuc_payload)

    return ket_qua


# ============================ SELENIUM LOGIN ============================

class MplisClient:
    def __init__(self, log_fn):
        self.log = log_fn
        self.session = None
        self.driver = None

    def open_browser_and_fill_login(self, username, password):
        options = Options()
        options.add_argument("--start-maximized")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.get(REFERER_URL)

        try:
            inputs = self.driver.find_elements(By.CSS_SELECTOR, "input")
            user_box = None
            pass_box = None
            for inp in inputs:
                typ = (inp.get_attribute("type") or "").lower()
                if not user_box and typ in ["text", "email"]:
                    user_box = inp
                if not pass_box and typ == "password":
                    pass_box = inp
            if user_box and pass_box:
                user_box.clear()
                user_box.send_keys(username)
                pass_box.clear()
                pass_box.send_keys(password)
                pass_box.send_keys(Keys.ENTER)
                self.log("Đã điền thông tin đăng nhập, chờ trang load...")
        except Exception as e:
            self.log(f"Không tự điền được form login ({e}), hãy đăng nhập tay trên Chrome.")

    def build_session_from_browser(self):
        if not self.driver:
            raise RuntimeError("Chưa mở trình duyệt.")

        token = lay_token_tu_trang(self.driver)
        if not token:
            raise RuntimeError("Không lấy được token. Kiểm tra đã đăng nhập và đang ở đúng trang chưa.")

        session = requests.Session()
        user_agent = self.driver.execute_script("return navigator.userAgent;")

        session.headers.update({
            "User-Agent": user_agent,
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": "https://dla.mplis.gov.vn",
            "Referer": REFERER_URL,
            "__requestverificationtoken": token,
            "__RequestVerificationToken": token,
            "RequestVerificationToken": token,
        })

        for c in self.driver.get_cookies():
            session.cookies.set(
                name=c["name"],
                value=c["value"],
                domain=c.get("domain"),
                path=c.get("path", "/"),
            )

        self.session = session
        self.log("Đã lấy session + token thành công.")

    def close_browser(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


# ============================ GUI ============================

ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")


class App:
    def __init__(self, root):
        self.root = root
        root.title("Cập nhật Chủ sở hữu (Tổ chức) MPLIS - hàng loạt")
        root.geometry("900x680")

        self.client = MplisClient(self.log)
        self.df = None
        self.running = False
        self.stop_flag = False

        frm = ctk.CTkFrame(root)
        frm.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(frm, text="Username:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.ent_user = ctk.CTkEntry(frm, width=220)
        self.ent_user.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        ctk.CTkLabel(frm, text="Password:").grid(row=0, column=2, sticky="w", padx=5, pady=5)
        self.ent_pass = ctk.CTkEntry(frm, width=220, show="*")
        self.ent_pass.grid(row=0, column=3, sticky="w", padx=5, pady=5)

        ctk.CTkLabel(frm, text="Tỉnh mặc định:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.ent_tinh = ctk.CTkEntry(frm, width=100)
        self.ent_tinh.insert(0, str(TINH_ID_MAC_DINH))
        self.ent_tinh.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(frm, text="(dùng khi Excel không có cột 'Tỉnh')").grid(
            row=1, column=2, columnspan=2, sticky="w", padx=5
        )

        ctk.CTkLabel(frm, text="File Excel:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.var_excel = tk.StringVar()
        ctk.CTkEntry(frm, textvariable=self.var_excel, width=460).grid(
            row=2, column=1, columnspan=2, sticky="we", padx=5, pady=5
        )
        ctk.CTkButton(frm, text="Chọn Excel...", command=self.chon_excel).grid(row=2, column=3, sticky="w", padx=5)

        self.var_gui_that = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            frm,
            text="Gửi update THẬT lên MPLIS (không tick = chỉ kiểm tra, không gửi)",
            variable=self.var_gui_that,
        ).grid(row=3, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 0))

        btn_frm = ctk.CTkFrame(root)
        btn_frm.pack(fill="x", padx=10)

        self.btn_login = ctk.CTkButton(btn_frm, text="1. Mở Chrome đăng nhập", command=self.mo_chrome)
        self.btn_login.pack(side="left", padx=5, pady=5)

        self.btn_confirm = ctk.CTkButton(
            btn_frm, text="2. Đã đăng nhập xong → Lấy session", command=self.lay_session, state="disabled"
        )
        self.btn_confirm.pack(side="left", padx=5, pady=5)

        self.btn_run = ctk.CTkButton(btn_frm, text="3. Chạy", command=self.chay, state="disabled")
        self.btn_run.pack(side="left", padx=5, pady=5)

        self.btn_stop = ctk.CTkButton(btn_frm, text="Dừng", command=self.dung, state="disabled")
        self.btn_stop.pack(side="left", padx=5, pady=5)

        self.progress = ctk.CTkProgressBar(root)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=10, pady=(10, 0))
        self.lbl_status = ctk.CTkLabel(root, text="Chưa chạy")
        self.lbl_status.pack(anchor="w", padx=10)

        self.txt = ctk.CTkTextbox(root, wrap="word")
        self.txt.pack(fill="both", expand=True, padx=10, pady=10)

        root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---------- UI helpers ----------
    def log(self, msg):
        def _append():
            self.txt.insert("end", f"{datetime.now().strftime('%H:%M:%S')}  {msg}\n")
            self.txt.see("end")
        self.root.after(0, _append)

    def set_status(self, msg):
        self.root.after(0, lambda: self.lbl_status.configure(text=msg))

    def chon_excel(self):
        f = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Tất cả", "*.*")],
        )
        if f:
            self.var_excel.set(f)

    # ---------- actions ----------
    def mo_chrome(self):
        username = self.ent_user.get().strip()
        password = self.ent_pass.get()
        if not username or not password:
            messagebox.showwarning("Thiếu thông tin", "Nhập username và password trước.")
            return

        self.btn_login.configure(state="disabled")

        def _work():
            try:
                self.log("Đang mở Chrome...")
                self.client.open_browser_and_fill_login(username, password)
                self.log("Chrome đã mở. Hoàn tất đăng nhập (OTP, captcha... nếu có) rồi bấm nút 2.")
                self.root.after(0, lambda: self.btn_confirm.configure(state="normal"))
            except Exception as e:
                self.log(f"Lỗi mở Chrome: {e}")
                self.root.after(0, lambda: self.btn_login.configure(state="normal"))

        threading.Thread(target=_work, daemon=True).start()

    def lay_session(self):
        def _work():
            try:
                self.client.build_session_from_browser()
                self.root.after(0, lambda: self.btn_run.configure(state="normal"))
            except Exception as e:
                self.log(f"{e}")

        threading.Thread(target=_work, daemon=True).start()

    def dung(self):
        self.stop_flag = True
        self.log("Đã yêu cầu dừng, sẽ dừng sau bản ghi hiện tại...")

    def doc_excel(self, path):
        df = pd.read_excel(path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        missing = [c for c in REQUIRED_COLS if c not in df.columns]
        if missing:
            raise RuntimeError(f"Excel thiếu cột: {', '.join(missing)}. Cần đủ: {', '.join(REQUIRED_COLS)}")

        for c in OPTIONAL_COLS:
            if c not in df.columns:
                df[c] = ""

        for c in df.columns:
            df[c] = df[c].map(clean_cell)

        df = df[df[COL_GCN] != ""].copy()
        return df

    def chay(self):
        if self.running:
            return

        excel = self.var_excel.get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showwarning("Thiếu Excel", "Chọn file Excel hợp lệ.")
            return

        try:
            tinh_id_mac_dinh = int(self.ent_tinh.get().strip() or TINH_ID_MAC_DINH)
        except ValueError:
            messagebox.showwarning("Sai Tỉnh", "Tỉnh mặc định phải là số.")
            return

        try:
            df = self.doc_excel(excel)
        except Exception as e:
            messagebox.showerror("Lỗi Excel", str(e))
            return

        if df.empty:
            messagebox.showwarning("Excel rỗng", "Không có bản ghi nào để xử lý (cột 'GCN' trống hết).")
            return

        try:
            o_dang_so = tim_o_dinh_dang_so(excel, COT_CO_THE_MAT_SO_0_DAU)
        except Exception as e:
            o_dang_so = []
            self.log(f"Không kiểm tra được định dạng ô (bỏ qua cảnh báo số 0 đầu): {e}")

        if o_dang_so:
            xem_truoc = "\n".join(
                f"  - Dòng {dong}, cột '{cot}': đang đọc được {gia_tri!r}"
                for dong, cot, gia_tri in o_dang_so[:15]
            )
            if len(o_dang_so) > 15:
                xem_truoc += f"\n  ... và {len(o_dang_so) - 15} ô khác."
            if not messagebox.askyesno(
                "Cảnh báo: ô đang ở định dạng Số",
                "Các ô sau đang được lưu dưới dạng SỐ trong file Excel, không phải Text:\n\n"
                f"{xem_truoc}\n\n"
                "Nếu giá trị thật có số 0 ở đầu (vd mã số doanh nghiệp \"0301234567\"), Excel đã "
                "làm rụng mất số 0 đó TRƯỚC KHI file được đọc - không có cách nào khôi phục lại "
                "được nữa, dữ liệu gửi lên MPLIS sẽ SAI. Cần sửa lại: định dạng cột là Text rồi "
                "nhập lại (hoặc gõ thêm dấu ' trước số), sau đó chọn lại file.\n\n"
                "Vẫn muốn tiếp tục với dữ liệu hiện tại (bỏ qua cảnh báo)?",
            ):
                return

        gui_that = self.var_gui_that.get()
        che_do = "GỬI UPDATE THẬT lên MPLIS" if gui_that else "CHỈ XEM TRƯỚC (không gửi update)"
        canh_bao = "\n\nThao tác GHI DỮ LIỆU THẬT, không tự hoàn tác được." if gui_that else ""
        if not messagebox.askyesno(
            "Xác nhận",
            f"Chế độ: {che_do}\nSẽ xử lý {len(df)} bản ghi.{canh_bao}\nTiếp tục?",
        ):
            return

        self.running = True
        self.stop_flag = False
        self.btn_run.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.progress.set(0)

        threading.Thread(
            target=self._run_batch,
            args=(df, excel, tinh_id_mac_dinh, gui_that),
            daemon=True,
        ).start()

    def _run_batch(self, df, excel_path, tinh_id_mac_dinh, gui_that):
        results = []
        total = len(df)

        for i, row in df.iterrows():
            if self.stop_flag:
                self.log("Đã dừng theo yêu cầu.")
                break

            sph = row[COL_GCN]
            tinh_id = int(row[COL_TINH]) if row.get(COL_TINH) else tinh_id_mac_dinh
            self.set_status(f"Đang xử lý {i + 1}/{total}: {sph}")
            self.log(f"[{i + 1}/{total}] GCN: {sph} | Tỉnh: {tinh_id}")

            ket_qua = "OK"
            loi = ""
            try:
                overrides, dai_dien_overrides, ngay_cap, so_giay_to = build_overrides_tu_dong(row)
                kq = xu_ly_1_dong(
                    self.client.session, sph, tinh_id, overrides, dai_dien_overrides,
                    ngay_cap, so_giay_to, gui_that,
                )

                for hau_to in ("ca_nhan_result", "to_chuc_result"):
                    ket = kq.get(hau_to)
                    if ket is not None and not ket.get("success", True):
                        ket_qua = "LỖI"
                        loi = f"Server trả success=false ({hau_to})"

                if ket_qua == "OK":
                    self.log(f"   → {'Đã gửi update' if gui_that else 'Đã kiểm tra, sẵn sàng gửi'}")
            except Exception as e:
                ket_qua = "LỖI"
                loi = str(e)
                self.log(f"   → Lỗi: {loi}")

            results.append({
                COL_GCN: sph,
                COL_TINH: tinh_id,
                "Kết quả": ket_qua,
                "Lỗi": loi,
            })

            self.root.after(0, lambda v=(i + 1) / total: self.progress.set(v))

        try:
            out_path = os.path.join(
                os.path.dirname(excel_path),
                f"ket_qua_cap_nhat_chusohuu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            pd.DataFrame(results).to_excel(out_path, index=False)
            self.log(f"Đã xuất file kết quả: {out_path}")
        except Exception as e:
            self.log(f"Không xuất được file kết quả: {e}")

        ok = sum(1 for r in results if r["Kết quả"] == "OK")
        loi = sum(1 for r in results if r["Kết quả"] == "LỖI")
        self.log(f"===== HOÀN TẤT: {ok} thành công | {loi} lỗi / tổng {len(results)} =====")
        self.set_status(f"Hoàn tất: {ok} OK | {loi} lỗi / {len(results)}")

        self.running = False
        self.root.after(0, lambda: (self.btn_run.configure(state="normal"), self.btn_stop.configure(state="disabled")))

    def on_close(self):
        if self.running and not messagebox.askyesno("Đang chạy", "Đang xử lý, thoát luôn?"):
            return
        self.client.close_browser()
        self.root.destroy()


def main():
    root = ctk.CTk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
