# -*- coding: utf-8 -*-
"""
Tool kiểm tra và xóa đơn đăng ký MPLIS theo danh sách tờ/thửa trong Excel.

Luồng xử lý:
1. Mở Chrome và đăng nhập MPLIS.
2. Lấy cookie + __RequestVerificationToken từ Chrome.
3. Đọc file Excel có cột soto và sothua.
4. Tra AdvancedSearchTinhHinhDangKy theo mã xã, số tờ, số thửa.
5. Tùy chế độ được chọn trên UI:
   - CHỈ KIỂM TRA: chỉ tra GCN + hồ sơ quét, không xóa.
   - KIỂM TRA RỒI XÓA: chỉ xóa khi đồng thời không có soPhatHanh
     trong ListGiayChungNhan và ListHoSoQuet rỗng.
   - XÓA THẲNG: xóa tất cả đơn tìm thấy, KHÔNG kiểm tra GCN/HSQ.
6. Ghi kết quả theo từng đơn vào Excel và tự lưu sau mỗi 5 thửa đất.

Cài thư viện:
    python -m pip install requests selenium webdriver-manager openpyxl

File Excel đầu vào:
- Định dạng .xlsx hoặc .xlsm.
- Có hàng tiêu đề chứa cột soto và sothua.
- Chương trình cũng nhận một số tên tương đương như: Số tờ, so_to,
  soHieuToBanDo, Số thửa, so_thua, soThuTuThua.
"""

from __future__ import annotations

import os
import threading
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


# ============================ CẤU HÌNH ============================

BASE_URL = "https://dla.mplis.gov.vn"

# Trang mở để đăng nhập và lấy token.
REFERER_LOGIN = f"{BASE_URL}/dc/DonDangKy/KeKhaiDangKyV2"
REFERER_API = REFERER_LOGIN

URL_TIM_DON = f"{BASE_URL}/dc/DangKyAjax/AdvancedSearchTinhHinhDangKy"
URL_CHI_TIET = f"{BASE_URL}/dc/DangKyAjax/GetThongTinDangKyByTinhHinhDangKyIds"
URL_XOA_DON = f"{BASE_URL}/dc/DangKyAjax/DeleteDonDangKyByTinhHinhDangKyId"

PAGE_SIZE = 10
TIMEOUT = 120
SAVE_EVERY_PARCELS = 5
REQUEST_DELAY_SECONDS = 0.15

# Các chế độ xử lý.
MODE_CHI_KIEM_TRA = "chi_kiem_tra"
MODE_KIEM_TRA_XOA = "kiem_tra_xoa"
MODE_XOA_THANG = "xoa_thang"

OUTPUT_HEADERS = [
    "stt",
    "soto",
    "sothua",
    "tinhhinhdangkyid",
    "sogcn",
    "hosoquet",
    "ketqua xuly",
]


# ============================ HELPER ============================


def lay_token_tu_trang(driver: webdriver.Chrome) -> str:
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    value = driver.execute_script(js)
    return str(value or "").strip()


def chuan_hoa_ten_cot(value: Any) -> str:
    """Chuẩn hóa tên cột để nhận dạng soto/sothua linh hoạt."""
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return "".join(char for char in text if char.isalnum())


def chuan_hoa_gia_tri_excel(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def rut_gon_text(value: Any, limit: int = 500) -> str:
    text = str(value or "").strip().replace("\r", " ").replace("\n", " ")
    if len(text) > limit:
        return text[:limit] + "..."
    return text


# ============================ ĐỌC FILE EXCEL ============================


def tim_cot_to_thua(worksheet) -> tuple[int, int, int]:
    """
    Tìm hàng tiêu đề và vị trí cột số tờ/số thửa trong 10 hàng đầu.
    Trả về: (header_row, so_to_column, so_thua_column)
    """
    aliases_so_to = {
        "soto",
        "sotobando",
        "sohieutobando",
        "tobando",
        "to",
    }
    aliases_so_thua = {
        "sothua",
        "sothututhua",
        "thuadat",
        "thua",
    }

    max_scan_row = min(10, worksheet.max_row)

    for row_index in range(1, max_scan_row + 1):
        so_to_column = None
        so_thua_column = None

        for column_index in range(1, worksheet.max_column + 1):
            name = chuan_hoa_ten_cot(
                worksheet.cell(row=row_index, column=column_index).value
            )

            if name in aliases_so_to and so_to_column is None:
                so_to_column = column_index

            if name in aliases_so_thua and so_thua_column is None:
                so_thua_column = column_index

        if so_to_column and so_thua_column:
            return row_index, so_to_column, so_thua_column

    raise ValueError(
        "Không tìm thấy đủ hai cột soto và sothua trong 10 hàng đầu của file Excel."
    )


def doc_danh_sach_thua(file_path: str) -> list[dict[str, Any]]:
    extension = Path(file_path).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("Chỉ hỗ trợ file Excel .xlsx hoặc .xlsm.")

    try:
        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True,
        )
    except BadZipFile as exc:
        raise ValueError(
            "File được chọn không phải file Excel .xlsx/.xlsm hợp lệ hoặc file đã bị hỏng."
        ) from exc

    try:
        worksheet = workbook.active
        header_row, so_to_column, so_thua_column = tim_cot_to_thua(worksheet)

        parcels: list[dict[str, Any]] = []

        for row_index in range(header_row + 1, worksheet.max_row + 1):
            so_to = chuan_hoa_gia_tri_excel(
                worksheet.cell(row=row_index, column=so_to_column).value
            )
            so_thua = chuan_hoa_gia_tri_excel(
                worksheet.cell(row=row_index, column=so_thua_column).value
            )

            # Bỏ qua hàng hoàn toàn trống.
            if not so_to and not so_thua:
                continue

            parcels.append(
                {
                    "excel_row": row_index,
                    "soto": so_to,
                    "sothua": so_thua,
                }
            )

        if not parcels:
            raise ValueError("File Excel không có dữ liệu tờ/thửa để xử lý.")

        return parcels
    finally:
        workbook.close()


# ============================ GHI FILE KẾT QUẢ ============================


class ExcelResultWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "KetQua"
        self.stt = 0

        self._setup_sheet()

    def _setup_sheet(self) -> None:
        self.worksheet.append(OUTPUT_HEADERS)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="B7B7B7")

        for cell in self.worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        widths = {
            1: 8,
            2: 12,
            3: 12,
            4: 22,
            5: 24,
            6: 14,
            7: 65,
        }

        for column_index, width in widths.items():
            self.worksheet.column_dimensions[get_column_letter(column_index)].width = width

        self.worksheet.freeze_panes = "A2"
        self.worksheet.auto_filter.ref = "A1:G1"
        self.worksheet.row_dimensions[1].height = 30

    def append_result(
        self,
        so_to: str,
        so_thua: str,
        tinh_hinh_dang_ky_id: Any,
        so_gcn: str,
        ho_so_quet: str,
        ket_qua_xu_ly: str,
    ) -> None:
        self.stt += 1
        self.worksheet.append(
            [
                self.stt,
                so_to,
                so_thua,
                tinh_hinh_dang_ky_id,
                so_gcn,
                ho_so_quet,
                ket_qua_xu_ly,
            ]
        )

        row_index = self.worksheet.max_row
        thin = Side(style="thin", color="D9D9D9")

        for cell in self.worksheet[row_index]:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        self.worksheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        self.worksheet.cell(row=row_index, column=6).alignment = Alignment(
            horizontal="center", vertical="top"
        )

        ket_qua_cell = self.worksheet.cell(row=row_index, column=7)
        ket_qua_lower = ket_qua_xu_ly.lower()

        if "xóa thành công" in ket_qua_lower:
            ket_qua_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            ket_qua_cell.font = Font(color="006100")
        elif "không xóa" in ket_qua_lower:
            ket_qua_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        elif "lỗi" in ket_qua_lower or "không xác định" in ket_qua_lower:
            ket_qua_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            ket_qua_cell.font = Font(color="9C0006")
        elif "đủ điều kiện" in ket_qua_lower:
            ket_qua_cell.fill = PatternFill("solid", fgColor="D9EAF7")

    def save(self) -> None:
        output_parent = Path(self.output_path).resolve().parent
        output_parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)

    def close(self) -> None:
        try:
            self.workbook.close()
        except Exception:
            pass


# ============================ CORE API ============================


class MplisClient:
    def __init__(self, log_fn):
        self.log = log_fn
        self.session: requests.Session | None = None
        self.driver: webdriver.Chrome | None = None

    # ---------- login ----------
    def open_browser_and_fill_login(self, username: str, password: str) -> None:
        options = Options()
        options.add_argument("--start-maximized")

        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.get(REFERER_LOGIN)
        time.sleep(2)

        try:
            inputs = self.driver.find_elements(By.CSS_SELECTOR, "input")
            user_box = None
            pass_box = None

            for inp in inputs:
                input_type = (inp.get_attribute("type") or "").lower()

                if user_box is None and input_type in {"text", "email"}:
                    user_box = inp

                if pass_box is None and input_type == "password":
                    pass_box = inp

            if user_box and pass_box:
                user_box.clear()
                user_box.send_keys(username)
                pass_box.clear()
                pass_box.send_keys(password)
                pass_box.send_keys(Keys.ENTER)
                self.log("Đã điền thông tin đăng nhập, chờ trang tải...")
            else:
                self.log("Không nhận dạng được form đăng nhập, hãy đăng nhập tay trên Chrome.")

        except Exception as exc:
            self.log(f"Không tự điền được form đăng nhập ({exc}), hãy đăng nhập tay.")

    def build_session_from_browser(self) -> None:
        if not self.driver:
            raise RuntimeError("Chưa mở trình duyệt.")

        token = lay_token_tu_trang(self.driver)

        if not token:
            raise RuntimeError(
                "Không lấy được token. Hãy bảo đảm đã đăng nhập thành công và đang mở trang MPLIS."
            )

        session = requests.Session()
        user_agent = self.driver.execute_script("return navigator.userAgent;")

        # Không đặt Content-Type chung vì API tìm kiếm gửi Form Data,
        # còn API chi tiết và xóa gửi JSON.
        session.headers.update(
            {
                "User-Agent": user_agent,
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Accept-Language": "vi-VN,vi;q=0.9",
                "X-Requested-With": "XMLHttpRequest",
                "Origin": BASE_URL,
                "Referer": REFERER_API,
                "__requestverificationtoken": token,
            }
        )

        for cookie in self.driver.get_cookies():
            session.cookies.set(
                name=cookie["name"],
                value=cookie["value"],
                domain=cookie.get("domain"),
                path=cookie.get("path", "/"),
            )

        self.session = session
        self.log("✅ Đã lấy session + token thành công.")

    def close_browser(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

    # ---------- request helpers ----------
    def _require_session(self) -> requests.Session:
        if self.session is None:
            raise RuntimeError("Chưa lấy session từ Chrome.")
        return self.session

    @staticmethod
    def _check_response(response: requests.Response, api_name: str) -> None:
        if response.status_code in {301, 302, 303, 307, 308} or response.headers.get(
            "Location"
        ):
            raise RuntimeError(
                f"{api_name} bị chuyển hướng (HTTP {response.status_code}). "
                "Phiên đăng nhập có thể đã hết hạn."
            )

        if response.status_code == 404:
            raise RuntimeError(f"{api_name}: endpoint không tồn tại (404).")

        response.raise_for_status()

    def _response_json(self, response: requests.Response, api_name: str) -> Any:
        self._check_response(response, api_name)

        try:
            return response.json()
        except requests.JSONDecodeError as exc:
            raise RuntimeError(
                f"{api_name} không trả JSON (HTTP {response.status_code}): "
                f"{response.text[:500]}"
            ) from exc

    # ---------- payload tìm kiếm ----------
    @staticmethod
    def build_search_payload(
        xa_id: str,
        so_to: str,
        so_thua: str,
        start: int = 0,
        length: int = PAGE_SIZE,
        draw: int = 1,
    ) -> dict[str, str]:
        payload = {
            "draw": str(draw),
            "order[0][column]": "5",
            "order[0][dir]": "desc",
            "start": str(start),
            "length": str(length),
            "search[value]": "",
            "search[regex]": "false",
            "model[xaId]": xa_id,
            "model[huyenId]": "",
            "model[tinhHinhDangKyId]": "",
            "model[maDon]": "",
            "model[soThuTu]": "",
            "model[ngayTiepNhan]": "",
            "model[thoiDiemDangKy]": "",
            "model[loaiGiayChungNhanId]": "",
            "model[soPhatHanh]": "",
            "model[maVach]": "",
            "model[soVaoSo]": "",
            "model[soVaoSoCu]": "",
            "model[ngayVaoSo]": "",
            "model[soHoSoGoc]": "",
            "model[soHoSoGocCu]": "",
            "model[hoTen]": "",
            "model[soGiayTo]": "",
            "model[namSinh]": "",
            "model[soThuTuThua]": so_thua,
            "model[soHieuToBanDo]": so_to,
            "model[soThuTuThuaCu]": "",
            "model[soHieuToBanDoCu]": "",
            "model[soNha]": "",
            "model[diaChiChiTiet]": "",
            "model[dieuKienCapGiay]": "",
            "model[phucHoiDuLieu]": "false",
        }

        columns = [
            ("", "", "true", "false"),
            ("tinhHinhDangKyId", "tinhHinhDangKyId", "true", "true"),
            ("maDon", "maDon", "true", "true"),
            ("soThuTu", "soThuTu", "true", "true"),
            ("DaiDienKhaiTrinh", "DaiDienKhaiTrinh", "true", "false"),
            ("ngayTiepNhan", "ngayTiepNhan", "true", "true"),
            ("thoiDiemDangKy", "thoiDiemDangKy", "true", "true"),
        ]

        for index, (data, name, searchable, orderable) in enumerate(columns):
            payload[f"columns[{index}][data]"] = data
            payload[f"columns[{index}][name]"] = name
            payload[f"columns[{index}][searchable]"] = searchable
            payload[f"columns[{index}][orderable]"] = orderable
            payload[f"columns[{index}][search][value]"] = ""
            payload[f"columns[{index}][search][regex]"] = "false"

        return payload

    # ---------- API ----------
    def tim_don_mot_trang(
        self,
        xa_id: str,
        so_to: str,
        so_thua: str,
        start: int,
        length: int,
        draw: int,
    ) -> tuple[list[dict[str, Any]], int | None]:
        session = self._require_session()
        payload = self.build_search_payload(
            xa_id=xa_id,
            so_to=so_to,
            so_thua=so_thua,
            start=start,
            length=length,
            draw=draw,
        )

        response = session.post(
            URL_TIM_DON,
            data=payload,
            timeout=TIMEOUT,
            allow_redirects=False,
        )

        result = self._response_json(response, "AdvancedSearchTinhHinhDangKy")
        rows = result.get("data") or []

        if not isinstance(rows, list):
            raise RuntimeError("API tìm đơn trả trường data không phải danh sách.")

        total_value = result.get("recordsFiltered")
        try:
            total = int(total_value)
        except (TypeError, ValueError):
            total = None

        return rows, total

    def tim_tat_ca_don(
        self,
        xa_id: str,
        so_to: str,
        so_thua: str,
    ) -> list[dict[str, Any]]:
        all_rows: list[dict[str, Any]] = []
        start = 0
        draw = 1

        while True:
            rows, total = self.tim_don_mot_trang(
                xa_id=xa_id,
                so_to=so_to,
                so_thua=so_thua,
                start=start,
                length=PAGE_SIZE,
                draw=draw,
            )

            if not rows:
                break

            all_rows.extend(rows)
            start += len(rows)
            draw += 1

            if total is not None and start >= total:
                break

            if len(rows) < PAGE_SIZE:
                break

        # Loại ID trùng nhưng giữ thứ tự xuất hiện.
        unique_rows: list[dict[str, Any]] = []
        seen_ids: set[str] = set()

        for row in all_rows:
            tinh_hinh_id = row.get("tinhHinhDangKyId")
            id_text = str(tinh_hinh_id or "").strip()

            if not id_text or id_text in seen_ids:
                continue

            seen_ids.add(id_text)
            unique_rows.append(row)

        return unique_rows

    def lay_chi_tiet_don(self, tinh_hinh_dang_ky_id: int) -> dict[str, Any] | None:
        session = self._require_session()
        payload = {
            "getHoSoQuet": True,
            "tinhHinhDangKyIds": [int(tinh_hinh_dang_ky_id)],
        }

        response = session.post(
            URL_CHI_TIET,
            json=payload,
            timeout=TIMEOUT,
            allow_redirects=False,
        )

        result = self._response_json(
            response,
            "GetThongTinDangKyByTinhHinhDangKyIds",
        )

        values = result.get("value") or []

        if isinstance(values, dict):
            values = [values]

        if not isinstance(values, list) or not values:
            return None

        for item in values:
            if not isinstance(item, dict):
                continue

            tinh_hinh = item.get("TinhHinhDangKy") or {}
            current_id = tinh_hinh.get("tinhHinhDangKyId")

            if str(current_id) == str(tinh_hinh_dang_ky_id):
                return item

        if len(values) == 1 and isinstance(values[0], dict):
            return values[0]

        return None

    def xoa_don(self, tinh_hinh_dang_ky_id: int) -> str:
        """
        Gửi đúng một lần lệnh xóa bằng JSON.
        Không tự retry để tránh nguy cơ gửi lệnh xóa hai lần.
        """
        session = self._require_session()
        payload = {"tinhHinhDangKyId": int(tinh_hinh_dang_ky_id)}

        response = session.post(
            URL_XOA_DON,
            json=payload,
            timeout=TIMEOUT,
            allow_redirects=False,
        )

        self._check_response(response, "DeleteDonDangKyByTinhHinhDangKyId")

        response_text = response.text.strip()
        parsed: Any = None

        if response_text:
            try:
                parsed = response.json()
            except requests.JSONDecodeError:
                parsed = response_text

        if parsed is False:
            raise RuntimeError("API trả về false.")

        if isinstance(parsed, dict):
            if parsed.get("success") is False:
                raise RuntimeError(rut_gon_text(parsed))

            message = (
                parsed.get("message")
                or parsed.get("Message")
                or parsed.get("msg")
                or parsed.get("error")
            )

            if message:
                return f"XÓA THÀNH CÔNG - {rut_gon_text(message)}"

            return f"XÓA THÀNH CÔNG - HTTP {response.status_code}"

        if parsed is True:
            return "XÓA THÀNH CÔNG - API trả về true"

        if response_text:
            return f"XÓA THÀNH CÔNG - HTTP {response.status_code}: {rut_gon_text(response_text)}"

        return f"XÓA THÀNH CÔNG - HTTP {response.status_code}"


# ============================ KIỂM TRA ĐIỀU KIỆN ============================


def danh_gia_chi_tiet(
    detail: dict[str, Any],
    check_gcn: bool = True,
    check_hsq: bool = True,
) -> dict[str, Any]:
    """
    Đánh giá điều kiện xóa theo các mục được chọn kiểm tra.

    - check_gcn=True: đơn có số phát hành GCN thì không xóa.
    - check_hsq=True: đơn có hồ sơ quét thì không xóa.
    - Chỉ trả du_dieu_kien_xoa=True khi TẤT CẢ mục được chọn đều "sạch".
    - Mục nào không được chọn thì bỏ qua, không ảnh hưởng kết quả.
    """
    if check_gcn and "ListGiayChungNhan" not in detail:
        return {
            "hop_le": False,
            "so_gcn": "",
            "ho_so_quet": "",
            "du_dieu_kien_xoa": False,
            "ly_do": "Phản hồi thiếu ListGiayChungNhan - không xóa.",
        }

    if check_hsq and "ListHoSoQuet" not in detail:
        return {
            "hop_le": False,
            "so_gcn": "",
            "ho_so_quet": "",
            "du_dieu_kien_xoa": False,
            "ly_do": "Phản hồi thiếu ListHoSoQuet - không xóa.",
        }

    list_gcn = detail.get("ListGiayChungNhan")
    list_ho_so_quet = detail.get("ListHoSoQuet")

    if list_gcn is None:
        list_gcn = []
    if list_ho_so_quet is None:
        list_ho_so_quet = []

    if check_gcn and not isinstance(list_gcn, list):
        return {
            "hop_le": False,
            "so_gcn": "",
            "ho_so_quet": "",
            "du_dieu_kien_xoa": False,
            "ly_do": "ListGiayChungNhan không phải danh sách - không xóa.",
        }

    if check_hsq and not isinstance(list_ho_so_quet, list):
        return {
            "hop_le": False,
            "so_gcn": "",
            "ho_so_quet": "",
            "du_dieu_kien_xoa": False,
            "ly_do": "ListHoSoQuet không phải danh sách - không xóa.",
        }

    if not isinstance(list_gcn, list):
        list_gcn = []
    if not isinstance(list_ho_so_quet, list):
        list_ho_so_quet = []

    so_phat_hanh_list: list[str] = []

    for gcn in list_gcn:
        if not isinstance(gcn, dict):
            continue

        so_phat_hanh = str(gcn.get("soPhatHanh") or "").strip()

        if so_phat_hanh and so_phat_hanh not in so_phat_hanh_list:
            so_phat_hanh_list.append(so_phat_hanh)

    co_so_gcn = bool(so_phat_hanh_list)

    # Bảo thủ: ListHoSoQuet chỉ cần có phần tử là xem như có hồ sơ quét.
    co_ho_so_quet = bool(list_ho_so_quet)

    so_gcn_text = "; ".join(so_phat_hanh_list)
    ho_so_quet_text = "có" if co_ho_so_quet else "không"

    # Chỉ mục nào được chọn kiểm tra mới có quyền chặn xóa.
    chan_vi_gcn = check_gcn and co_so_gcn
    chan_vi_hsq = check_hsq and co_ho_so_quet

    du_dieu_kien_xoa = not (chan_vi_gcn or chan_vi_hsq)

    if chan_vi_gcn and chan_vi_hsq:
        ly_do = "KHÔNG XÓA - Có số GCN và có hồ sơ quét."
    elif chan_vi_gcn:
        ly_do = "KHÔNG XÓA - Có số GCN."
    elif chan_vi_hsq:
        ly_do = "KHÔNG XÓA - Có hồ sơ quét."
    else:
        checked_parts = []
        if check_gcn:
            checked_parts.append("không có số GCN")
        if check_hsq:
            checked_parts.append("không có hồ sơ quét")
        ly_do = (
            "ĐỦ ĐIỀU KIỆN XÓA - " + " và ".join(checked_parts).capitalize() + "."
        )

    return {
        "hop_le": True,
        "so_gcn": so_gcn_text,
        "ho_so_quet": ho_so_quet_text,
        "du_dieu_kien_xoa": du_dieu_kien_xoa,
        "ly_do": ly_do,
    }


# ============================ TKINTER UI ============================


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("Xóa đơn đăng ký MPLIS")
        root.geometry("980x760")
        root.minsize(900, 680)

        self.client = MplisClient(self.log)
        self.running = False
        self.stop_flag = False

        self.var_input_file = tk.StringVar()
        self.var_output_file = tk.StringVar()
        self.var_mode = tk.StringVar(value=MODE_CHI_KIEM_TRA)
        self.var_check_gcn = tk.BooleanVar(value=True)
        self.var_check_hsq = tk.BooleanVar(value=True)

        self._build_ui()
        root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self) -> None:
        login_frame = ttk.LabelFrame(self.root, text="Đăng nhập MPLIS", padding=10)
        login_frame.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Label(login_frame, text="Username:").grid(row=0, column=0, sticky="w")
        self.ent_user = ttk.Entry(login_frame, width=32)
        self.ent_user.grid(row=0, column=1, sticky="ew", padx=5, pady=3)

        ttk.Label(login_frame, text="Password:").grid(row=0, column=2, sticky="w")
        self.ent_pass = ttk.Entry(login_frame, width=32, show="*")
        self.ent_pass.grid(row=0, column=3, sticky="ew", padx=5, pady=3)

        ttk.Label(login_frame, text="Mã xã (xaId):").grid(row=1, column=0, sticky="w")
        self.ent_xa_id = ttk.Entry(login_frame, width=20)
        self.ent_xa_id.grid(row=1, column=1, sticky="w", padx=5, pady=3)

        login_frame.columnconfigure(1, weight=1)
        login_frame.columnconfigure(3, weight=1)

        file_frame = ttk.LabelFrame(self.root, text="File Excel", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(file_frame, text="File đầu vào:").grid(row=0, column=0, sticky="w")
        self.ent_input_file = ttk.Entry(
            file_frame,
            textvariable=self.var_input_file,
        )
        self.ent_input_file.grid(row=0, column=1, sticky="ew", padx=5, pady=3)
        ttk.Button(
            file_frame,
            text="Duyệt file...",
            command=self.chon_file_input,
        ).grid(row=0, column=2, padx=5, pady=3)

        ttk.Label(file_frame, text="File kết quả:").grid(row=1, column=0, sticky="w")
        self.ent_output_file = ttk.Entry(
            file_frame,
            textvariable=self.var_output_file,
        )
        self.ent_output_file.grid(row=1, column=1, sticky="ew", padx=5, pady=3)
        ttk.Button(
            file_frame,
            text="Chọn nơi lưu...",
            command=self.chon_file_output,
        ).grid(row=1, column=2, padx=5, pady=3)

        file_frame.columnconfigure(1, weight=1)

        mode_frame = ttk.LabelFrame(self.root, text="Chế độ xử lý", padding=8)
        mode_frame.pack(fill="x", padx=10, pady=5)

        self.rad_chi_kiem_tra = ttk.Radiobutton(
            mode_frame,
            text="Chỉ kiểm tra GCN, hồ sơ quét (KHÔNG xóa)",
            variable=self.var_mode,
            value=MODE_CHI_KIEM_TRA,
        )
        self.rad_chi_kiem_tra.pack(anchor="w", padx=5, pady=1)

        self.rad_kiem_tra_xoa = ttk.Radiobutton(
            mode_frame,
            text="Kiểm tra GCN, hồ sơ quét → chỉ xóa đơn đủ điều kiện",
            variable=self.var_mode,
            value=MODE_KIEM_TRA_XOA,
        )
        self.rad_kiem_tra_xoa.pack(anchor="w", padx=5, pady=1)

        # Chọn mục cần kiểm tra (áp dụng cho 2 chế độ có kiểm tra ở trên).
        check_frame = ttk.Frame(mode_frame)
        check_frame.pack(anchor="w", padx=28, pady=(0, 4))

        ttk.Label(check_frame, text="Điều kiện kiểm tra:").pack(side="left")

        self.chk_gcn = ttk.Checkbutton(
            check_frame,
            text="Số GCN",
            variable=self.var_check_gcn,
        )
        self.chk_gcn.pack(side="left", padx=8)

        self.chk_hsq = ttk.Checkbutton(
            check_frame,
            text="Hồ sơ quét",
            variable=self.var_check_hsq,
        )
        self.chk_hsq.pack(side="left", padx=8)

        ttk.Label(
            check_frame,
            text="(tích mục nào thì đơn có mục đó sẽ KHÔNG bị xóa)",
            foreground="gray",
        ).pack(side="left", padx=5)

        self.rad_xoa_thang = ttk.Radiobutton(
            mode_frame,
            text="⚠ XÓA THẲNG - xóa mọi đơn tìm thấy, KHÔNG kiểm tra GCN, hồ sơ quét",
            variable=self.var_mode,
            value=MODE_XOA_THANG,
        )
        self.rad_xoa_thang.pack(anchor="w", padx=5, pady=1)

        self.mode_radios = [
            self.rad_chi_kiem_tra,
            self.rad_kiem_tra_xoa,
            self.rad_xoa_thang,
        ]

        ttk.Label(
            mode_frame,
            text="Tự lưu kết quả sau mỗi 5 thửa đất",
            foreground="blue",
        ).pack(anchor="e", padx=5)

        button_frame = ttk.Frame(self.root, padding=(10, 5))
        button_frame.pack(fill="x")

        self.btn_login = ttk.Button(
            button_frame,
            text="1. Mở Chrome đăng nhập",
            command=self.mo_chrome,
        )
        self.btn_login.pack(side="left", padx=5)

        self.btn_session = ttk.Button(
            button_frame,
            text="2. Đã đăng nhập xong → Lấy session",
            command=self.lay_session,
            state="disabled",
        )
        self.btn_session.pack(side="left", padx=5)

        self.btn_run = ttk.Button(
            button_frame,
            text="3. Bắt đầu xử lý Excel",
            command=self.chay,
            state="disabled",
        )
        self.btn_run.pack(side="left", padx=5)

        self.btn_stop = ttk.Button(
            button_frame,
            text="Dừng",
            command=self.dung,
            state="disabled",
        )
        self.btn_stop.pack(side="left", padx=5)

        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=(5, 0))

        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill="x", padx=10, pady=(3, 0))

        self.lbl_status = ttk.Label(status_frame, text="Chưa chạy")
        self.lbl_status.pack(side="left")

        self.lbl_count = ttk.Label(status_frame, text="", foreground="blue")
        self.lbl_count.pack(side="right")

        log_frame = ttk.LabelFrame(self.root, text="Nhật ký xử lý", padding=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=8)

        self.txt_log = tk.Text(log_frame, wrap="word", height=25)
        self.txt_log.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(log_frame, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=scrollbar.set)

    # ---------- UI helpers ----------
    def log(self, message: str) -> None:
        def append() -> None:
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.txt_log.insert("end", f"{timestamp}  {message}\n")
            self.txt_log.see("end")

        self.root.after(0, append)

    def set_status(self, message: str) -> None:
        self.root.after(0, lambda: self.lbl_status.config(text=message))

    def set_count(self, processed: int, total: int) -> None:
        self.root.after(
            0,
            lambda: self.lbl_count.config(text=f"Thửa: {processed} / {total}"),
        )

    def set_progress(self, value: int, maximum: int) -> None:
        def update() -> None:
            self.progress.config(maximum=max(maximum, 1), value=value)

        self.root.after(0, update)

    def set_mode_radios_state(self, state: str) -> None:
        for widget in [*self.mode_radios, self.chk_gcn, self.chk_hsq]:
            widget.config(state=state)

    # ---------- file actions ----------
    def chon_file_input(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel có cột soto, sothua",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled", "*.xlsm"),
                ("Tất cả file", "*.*"),
            ],
        )

        if not file_path:
            return

        self.var_input_file.set(file_path)

        input_path = Path(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = input_path.with_name(
            f"ket_qua_xoa_don_{timestamp}.xlsx"
        )
        self.var_output_file.set(str(output_path))

    def chon_file_output(self) -> None:
        initial_name = Path(self.var_output_file.get() or "ket_qua_xoa_don.xlsx").name

        file_path = filedialog.asksaveasfilename(
            title="Chọn nơi lưu file kết quả",
            defaultextension=".xlsx",
            initialfile=initial_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if file_path:
            self.var_output_file.set(file_path)

    # ---------- login actions ----------
    def mo_chrome(self) -> None:
        username = self.ent_user.get().strip()
        password = self.ent_pass.get()

        if not username or not password:
            messagebox.showwarning(
                "Thiếu thông tin",
                "Nhập username và password trước.",
            )
            return

        self.btn_login.config(state="disabled")

        def work() -> None:
            try:
                self.log("Đang mở Chrome...")
                self.client.open_browser_and_fill_login(username, password)
                self.log(
                    "Chrome đã mở. Hoàn tất đăng nhập/OTP nếu có, "
                    "sau đó bấm nút 2 để lấy session."
                )
                self.root.after(0, lambda: self.btn_session.config(state="normal"))
            except Exception as exc:
                self.log(f"❌ Lỗi mở Chrome: {exc}")
                self.root.after(0, lambda: self.btn_login.config(state="normal"))

        threading.Thread(target=work, daemon=True).start()

    def lay_session(self) -> None:
        self.btn_session.config(state="disabled")

        def work() -> None:
            try:
                self.client.build_session_from_browser()

                # Đã sao chép cookie và token sang requests.Session,
                # có thể đóng Chrome.
                self.client.close_browser()
                self.log("✅ Đã đóng Chrome sau khi lấy session.")

                self.root.after(
                    0,
                    lambda: self.btn_run.config(state="normal")
                )

            except Exception as exc:
                self.log(f"❌ {exc}")
                self.root.after(
                    0,
                    lambda: self.btn_session.config(state="normal")
                )

        threading.Thread(target=work, daemon=True).start()

    # ---------- processing actions ----------
    def dung(self) -> None:
        self.stop_flag = True
        self.log("⏸ Đã yêu cầu dừng. Sẽ dừng sau request hiện tại và lưu kết quả đã có.")

    def chay(self) -> None:
        if self.running:
            return

        xa_id = self.ent_xa_id.get().strip()
        input_file = self.var_input_file.get().strip()
        output_file = self.var_output_file.get().strip()
        mode = self.var_mode.get()
        check_gcn = self.var_check_gcn.get()
        check_hsq = self.var_check_hsq.get()

        if mode != MODE_XOA_THANG and not (check_gcn or check_hsq):
            messagebox.showwarning(
                "Thiếu điều kiện kiểm tra",
                "Chế độ có kiểm tra phải tích ít nhất một mục: Số GCN hoặc Hồ sơ quét.\n"
                "Nếu muốn xóa không kiểm tra gì, hãy chọn chế độ XÓA THẲNG.",
            )
            return

        if not xa_id:
            messagebox.showwarning("Thiếu thông tin", "Nhập mã xã (xaId).")
            return

        if not xa_id.isdigit():
            messagebox.showwarning("Sai mã xã", "Mã xã phải là số nguyên.")
            return

        if not input_file:
            messagebox.showwarning("Thiếu file", "Chọn file Excel đầu vào.")
            return

        if not os.path.isfile(input_file):
            messagebox.showwarning("Không tìm thấy file", input_file)
            return

        if not output_file:
            messagebox.showwarning("Thiếu file", "Chọn đường dẫn file kết quả.")
            return

        if Path(input_file).resolve() == Path(output_file).resolve():
            messagebox.showwarning(
                "Sai đường dẫn",
                "File kết quả không được trùng với file Excel đầu vào.",
            )
            return

        try:
            parcels = doc_danh_sach_thua(input_file)
        except Exception as exc:
            messagebox.showerror("Lỗi đọc Excel", str(exc))
            return

        if mode == MODE_XOA_THANG:
            confirm_text = (
                f"File có {len(parcels)} thửa đất.\n\n"
                "⚠ CẢNH BÁO: Chế độ XÓA THẲNG.\n\n"
                "Chương trình sẽ XÓA TẤT CẢ đơn đăng ký tìm thấy\n"
                "theo danh sách tờ/thửa, KHÔNG kiểm tra số GCN\n"
                "và KHÔNG kiểm tra hồ sơ quét.\n\n"
                "Tiếp tục?"
            )
            title = "CẢNH BÁO - Xóa thẳng không kiểm tra"
        else:
            dieu_kien_lines = []
            if check_gcn:
                dieu_kien_lines.append("- Không có số phát hành GCN")
            if check_hsq:
                dieu_kien_lines.append("- Không có hồ sơ quét")
            dieu_kien_text = "\n".join(dieu_kien_lines)

            if mode == MODE_KIEM_TRA_XOA:
                confirm_text = (
                    f"File có {len(parcels)} thửa đất.\n\n"
                    "Chương trình sẽ THỰC HIỆN XÓA các đơn thỏa TẤT CẢ điều kiện:\n"
                    f"{dieu_kien_text}\n\n"
                    "Đơn vi phạm bất kỳ điều kiện nào ở trên sẽ KHÔNG bị xóa.\n"
                    "Tiếp tục?"
                )
                title = "Xác nhận kiểm tra rồi xóa"
            else:
                confirm_text = (
                    f"File có {len(parcels)} thửa đất.\n\n"
                    "Chương trình đang ở chế độ CHỈ KIỂM TRA, KHÔNG XÓA.\n"
                    "Điều kiện được kiểm tra:\n"
                    f"{dieu_kien_text}\n\n"
                    "Tiếp tục?"
                )
                title = "Xác nhận kiểm tra"

        if not messagebox.askyesno(title, confirm_text):
            return

        # Xóa thẳng thì hỏi lại lần cuối để tránh bấm nhầm.
        if mode == MODE_XOA_THANG:
            if not messagebox.askyesno(
                "Xác nhận lần cuối",
                "Bạn CHẮC CHẮN muốn XÓA THẲNG không kiểm tra GCN, hồ sơ quét?\n"
                "Hành động này không thể hoàn tác.",
                icon="warning",
            ):
                return

        self.running = True
        self.stop_flag = False
        self.btn_run.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.set_mode_radios_state("disabled")
        self.progress.config(value=0, maximum=max(len(parcels), 1))
        self.set_count(0, len(parcels))

        threading.Thread(
            target=self._run_process,
            args=(xa_id, parcels, output_file, mode, check_gcn, check_hsq),
            daemon=True,
        ).start()

    def _run_process(
        self,
        xa_id: str,
        parcels: list[dict[str, Any]],
        output_file: str,
        mode: str,
        check_gcn: bool,
        check_hsq: bool,
    ) -> None:
        writer = ExcelResultWriter(output_file)
        processed_parcels = 0
        total_results = 0
        delete_success = 0
        delete_errors = 0
        skipped = 0

        try:
            dieu_kien_parts = []
            if check_gcn:
                dieu_kien_parts.append("GCN")
            if check_hsq:
                dieu_kien_parts.append("HSQ")
            dieu_kien_text = " + ".join(dieu_kien_parts)

            if mode == MODE_XOA_THANG:
                mode_text = "XÓA THẲNG (KHÔNG KIỂM TRA GCN, HSQ)"
            elif mode == MODE_KIEM_TRA_XOA:
                mode_text = f"KIỂM TRA {dieu_kien_text} RỒI XÓA ĐƠN ĐỦ ĐIỀU KIỆN"
            else:
                mode_text = f"CHỈ KIỂM TRA ({dieu_kien_text})"

            self.log("=" * 70)
            self.log(f"BẮT ĐẦU - Chế độ: {mode_text}")
            self.log(f"Mã xã: {xa_id}")
            self.log(f"Số thửa đầu vào: {len(parcels)}")
            self.log(f"File kết quả: {output_file}")

            for parcel_index, parcel in enumerate(parcels, start=1):
                if self.stop_flag:
                    self.log("⏹ Đã dừng theo yêu cầu.")
                    break

                so_to = parcel["soto"]
                so_thua = parcel["sothua"]
                excel_row = parcel["excel_row"]

                self.set_status(
                    f"Đang xử lý tờ {so_to or '?'} - thửa {so_thua or '?'}"
                )
                self.log(
                    f"--- [{parcel_index}/{len(parcels)}] "
                    f"Excel dòng {excel_row}: tờ {so_to}, thửa {so_thua} ---"
                )

                if not so_to or not so_thua:
                    writer.append_result(
                        so_to=so_to,
                        so_thua=so_thua,
                        tinh_hinh_dang_ky_id="",
                        so_gcn="",
                        ho_so_quet="",
                        ket_qua_xu_ly="BỎ QUA - Thiếu số tờ hoặc số thửa.",
                    )
                    total_results += 1
                    skipped += 1
                else:
                    try:
                        registrations = self.client.tim_tat_ca_don(
                            xa_id=xa_id,
                            so_to=so_to,
                            so_thua=so_thua,
                        )
                    except Exception as exc:
                        writer.append_result(
                            so_to=so_to,
                            so_thua=so_thua,
                            tinh_hinh_dang_ky_id="",
                            so_gcn="",
                            ho_so_quet="",
                            ket_qua_xu_ly=f"LỖI TRA CỨU - {rut_gon_text(exc)}",
                        )
                        total_results += 1
                        skipped += 1
                        self.log(f"   ❌ Lỗi tra cứu: {exc}")
                        registrations = []

                    if not registrations:
                        # Nếu trước đó là lỗi, đã ghi một dòng lỗi rồi.
                        # Chỉ ghi không tìm thấy khi thật sự request thành công và trả rỗng.
                        last_result = writer.worksheet.cell(
                            row=writer.worksheet.max_row,
                            column=7,
                        ).value
                        current_last_to = writer.worksheet.cell(
                            row=writer.worksheet.max_row,
                            column=2,
                        ).value
                        current_last_thua = writer.worksheet.cell(
                            row=writer.worksheet.max_row,
                            column=3,
                        ).value

                        already_error = (
                            str(current_last_to or "") == so_to
                            and str(current_last_thua or "") == so_thua
                            and str(last_result or "").startswith("LỖI TRA CỨU")
                        )

                        if not already_error:
                            writer.append_result(
                                so_to=so_to,
                                so_thua=so_thua,
                                tinh_hinh_dang_ky_id="",
                                so_gcn="",
                                ho_so_quet="",
                                ket_qua_xu_ly="KHÔNG TÌM THẤY ĐƠN ĐĂNG KÝ.",
                            )
                            total_results += 1
                            skipped += 1
                            self.log("   Không tìm thấy đơn đăng ký.")
                    else:
                        self.log(f"   Tìm thấy {len(registrations)} đơn đăng ký.")

                        for registration in registrations:
                            if self.stop_flag:
                                break

                            tinh_hinh_id = registration.get("tinhHinhDangKyId")

                            try:
                                tinh_hinh_id_int = int(tinh_hinh_id)
                            except (TypeError, ValueError):
                                writer.append_result(
                                    so_to=so_to,
                                    so_thua=so_thua,
                                    tinh_hinh_dang_ky_id=tinh_hinh_id or "",
                                    so_gcn="",
                                    ho_so_quet="",
                                    ket_qua_xu_ly="KHÔNG XÓA - tinhHinhDangKyId không hợp lệ.",
                                )
                                total_results += 1
                                skipped += 1
                                continue

                            # ---------- CHẾ ĐỘ XÓA THẲNG ----------
                            if mode == MODE_XOA_THANG:
                                self.log(f"   → XÓA THẲNG ID {tinh_hinh_id_int}")

                                try:
                                    result_text = self.client.xoa_don(
                                        tinh_hinh_id_int
                                    )
                                    delete_success += 1
                                except Exception as exc:
                                    result_text = (
                                        f"XÓA LỖI - {rut_gon_text(exc)}"
                                    )
                                    delete_errors += 1

                                writer.append_result(
                                    so_to=so_to,
                                    so_thua=so_thua,
                                    tinh_hinh_dang_ky_id=tinh_hinh_id_int,
                                    so_gcn="",
                                    ho_so_quet="không kiểm tra",
                                    ket_qua_xu_ly=result_text,
                                )
                                total_results += 1
                                self.log(f"      {result_text}")

                                time.sleep(REQUEST_DELAY_SECONDS)
                                continue

                            # ---------- CHẾ ĐỘ CÓ KIỂM TRA ----------
                            self.log(f"   → Kiểm tra ID {tinh_hinh_id_int}")

                            try:
                                detail = self.client.lay_chi_tiet_don(
                                    tinh_hinh_id_int
                                )
                            except Exception as exc:
                                writer.append_result(
                                    so_to=so_to,
                                    so_thua=so_thua,
                                    tinh_hinh_dang_ky_id=tinh_hinh_id_int,
                                    so_gcn="",
                                    ho_so_quet="",
                                    ket_qua_xu_ly=(
                                        "KHÔNG XÓA - Lỗi lấy chi tiết: "
                                        f"{rut_gon_text(exc)}"
                                    ),
                                )
                                total_results += 1
                                skipped += 1
                                self.log(f"      ❌ Lỗi lấy chi tiết: {exc}")
                                continue

                            if detail is None:
                                writer.append_result(
                                    so_to=so_to,
                                    so_thua=so_thua,
                                    tinh_hinh_dang_ky_id=tinh_hinh_id_int,
                                    so_gcn="",
                                    ho_so_quet="",
                                    ket_qua_xu_ly=(
                                        "KHÔNG XÓA - API không trả dữ liệu chi tiết."
                                    ),
                                )
                                total_results += 1
                                skipped += 1
                                self.log("      ⚠ API không trả dữ liệu chi tiết.")
                                continue

                            evaluation = danh_gia_chi_tiet(
                                detail,
                                check_gcn=check_gcn,
                                check_hsq=check_hsq,
                            )
                            so_gcn = evaluation["so_gcn"]
                            ho_so_quet = evaluation["ho_so_quet"]

                            if not evaluation["hop_le"]:
                                result_text = f"KHÔNG XÓA - {evaluation['ly_do']}"
                                skipped += 1

                            elif not evaluation["du_dieu_kien_xoa"]:
                                result_text = evaluation["ly_do"]
                                skipped += 1

                            elif mode == MODE_CHI_KIEM_TRA:
                                result_text = (
                                    "ĐỦ ĐIỀU KIỆN XÓA - CHƯA XÓA "
                                    "(chế độ chỉ kiểm tra)."
                                )

                            else:
                                try:
                                    result_text = self.client.xoa_don(
                                        tinh_hinh_id_int
                                    )
                                    delete_success += 1
                                except Exception as exc:
                                    result_text = (
                                        "ĐỦ ĐIỀU KIỆN NHƯNG XÓA LỖI - "
                                        f"{rut_gon_text(exc)}"
                                    )
                                    delete_errors += 1

                            writer.append_result(
                                so_to=so_to,
                                so_thua=so_thua,
                                tinh_hinh_dang_ky_id=tinh_hinh_id_int,
                                so_gcn=so_gcn,
                                ho_so_quet=ho_so_quet,
                                ket_qua_xu_ly=result_text,
                            )
                            total_results += 1

                            self.log(
                                f"      GCN: {so_gcn or 'không có'} | "
                                f"Hồ sơ quét: {ho_so_quet or 'không xác định'} | "
                                f"{result_text}"
                            )

                            time.sleep(REQUEST_DELAY_SECONDS)

                processed_parcels += 1
                self.set_progress(processed_parcels, len(parcels))
                self.set_count(processed_parcels, len(parcels))

                if processed_parcels % SAVE_EVERY_PARCELS == 0:
                    writer.save()
                    self.log(
                        f"💾 Đã tự lưu kết quả sau {processed_parcels} thửa đất."
                    )

            writer.save()
            self.log("💾 Đã lưu file kết quả cuối cùng.")

            self.log("=" * 70)
            self.log(f"Đã xử lý thửa: {processed_parcels}/{len(parcels)}")
            self.log(f"Tổng dòng kết quả: {total_results}")
            self.log(f"Xóa thành công: {delete_success}")
            self.log(f"Xóa lỗi: {delete_errors}")
            self.log(f"Không xóa/bỏ qua/lỗi tra cứu: {skipped}")
            self.log(f"File kết quả: {output_file}")

            if self.stop_flag:
                self.set_status(
                    f"Đã dừng: xử lý {processed_parcels}/{len(parcels)} thửa, đã lưu kết quả"
                )
            else:
                self.set_status(
                    f"Hoàn tất: {processed_parcels} thửa, {delete_success} đơn xóa thành công"
                )

        except PermissionError:
            self.log(
                "❌ Không lưu được file kết quả. Có thể file đang được mở trong Excel. "
                "Hãy đóng file rồi chạy lại."
            )
            self.set_status("Lỗi lưu file kết quả")

        except Exception as exc:
            self.log(f"❌ Lỗi chương trình: {exc}")
            self.set_status("Chương trình gặp lỗi")

            try:
                writer.save()
                self.log("💾 Đã cố gắng lưu phần kết quả đang có.")
            except Exception as save_exc:
                self.log(f"❌ Không lưu được kết quả dở dang: {save_exc}")

        finally:
            writer.close()
            self.running = False

            def reset_buttons() -> None:
                self.btn_run.config(state="normal")
                self.btn_stop.config(state="disabled")
                self.set_mode_radios_state("normal")

            self.root.after(0, reset_buttons)

    def on_close(self) -> None:
        if self.running:
            confirm = messagebox.askyesno(
                "Đang xử lý",
                "Chương trình đang chạy. Thoát ngay có thể làm mất tối đa 4 thửa chưa kịp lưu. Thoát?",
            )
            if not confirm:
                return

        self.stop_flag = True
        self.client.close_browser()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()